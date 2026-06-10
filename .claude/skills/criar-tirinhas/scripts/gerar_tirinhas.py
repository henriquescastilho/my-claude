#!/usr/bin/env python3
"""
Criar Tirinhas — L.COSTA (3S-001) — VERSÃO CONSOLIDADA
1 único xlsx por mês com todos os médicos como blocos. Cada bloco mantém:
header colorido + linhas das cirurgias + separador + linha TOTAL.
Entre blocos: 1 linha em branco.

Saída: caminho passado em --output
       (convenção: <pasta_da_empresa>/Tirinhas_LCOSTA_<Mes>_<Ano>.xlsx)

Regras de Taxa (L.COSTA):
- Padrão: 20% sobre TOTAL (col K) — valor cheio, sem subtrair DESCONTO
- URGÊNCIA / EMERGÊNCIA: -10% silencioso adicional ANTES da Taxa
- Override: Juliano Berteli de Figueiredo — Coluna 25% (Crânio default 20%)
- Override: Edson Vargas de Oliveira Netto — Coluna 25% (Crânio default 20%)
"""

import argparse, os, re
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    "header_bg":"FFE082","row_bg":"FFFDE7","val_bg":"FFFFFF",
    "total_bg":"FFF9C4","sep_bg":"FFF3E0",
    "txt_dark":"1A1A1A","txt_head":"333333","txt_dim":"666666",
    "txt_gold":"8B6914","border":"BDBDBD","border_hdr":"9E9E9E",
}

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def fnt(hex_, bold=False, size=12): return Font(color=hex_, bold=bold, size=size, name="Calibri")
def aln(h="left", v="center"): return Alignment(horizontal=h, vertical=v)
def border_all(color="BDBDBD"):
    s = Side(style="thin", color=color); return Border(left=s, right=s, top=s, bottom=s)

HOSPITAL_MAP = {
    "CASA DE PORTUGAL":"Casa de Portugal","CASA SAUDE S JOSE HUMAITA":"Casa de Saúde São José Humaitá",
    "CLINICA SAO GONCALO":"Clínica São Gonçalo","CLINICA SAO VICENTE":"Clínica São Vicente",
    "GLORIA D OR":"Glória D'Or","GLORIA DOR":"Glória D'Or",
    "HOSP BADIM":"Hospital Badim","HOSP BANGU":"Hospital Bangu",
    "HOSP BARRA DOR II NOVO":"Hospital Barra D'Or II Novo","HOSP CAXIAS DOR":"Hospital Caxias D'Or",
    "HOSP CHN":"Hospital CHN","HOSP COPA D OR":"Hospital Copa D'Or",
    "HOSP COPA STAR":"Hospital Copa Star",
    "HOSP DE FORCA AEREA DO GALEAO":"Hospital de Força Aérea do Galeão",
    "HOSP NORTE DOR CASCADURA":"Hospital Norte D'Or Cascadura","HOSP OESTE DOR":"Hospital Oeste D'Or",
    "HOSP PASTEUR":"Hospital Pasteur","HOSP PRO CARDIACO":"Hospital Pró-Cardíaco",
    "HOSP QUINTA DOR":"Hospital Quinta D'Or","HOSP RIOS DOR":"Hospital Rios D'Or",
    "HOSP SAMARITANO":"Hospital Samaritano","HOSP SAMARITANO - BOTAFOGO":"Hospital Samaritano - Botafogo",
    "HOSP SAO LUCAS-COPACABANA":"Hospital São Lucas - Copacabana",
    "HOSP UNIMED ARARUAMA":"Hospital Unimed Araruama","HOSP UNIMED BARRA":"Hospital Unimed Barra",
    "HOSP VITORIA RJ":"Hospital Vitória - RJ","NITEROI D OR":"Niterói D'Or","PRONTONIL":"Prontonil",
}
WORD_FIX = {
    "HOSP":"Hospital","MAT":"Maternidade","SAO":"São","STA":"Santa","STO":"Santo","SRA":"Senhora",
    "NSR":"Nossa Senhora","ADAO":"Adão","ANTONIO":"Antônio","JOAO":"João","JOSE":"José",
    "JULIO":"Júlio","MARIO":"Mário","OTAVIO":"Otávio","OCTAVIO":"Otávio","ROGERIO":"Rogério",
    "SAVIO":"Sávio","VINICIUS":"Vinícius","CLAUDIO":"Cláudio","MARCIO":"Márcio","FLAVIO":"Flávio",
    "FABIO":"Fábio","LUCIO":"Lúcio","CAIO":"Caio","AURELIO":"Aurélio","HELIO":"Hélio",
    "SERGIO":"Sérgio","CICERO":"Cícero","HENRIQUE":"Henrique","CORREA":"Corrêa","SIMOES":"Simões",
    "GUIMARAES":"Guimarães","GONCALVES":"Gonçalves","MOURAO":"Mourão","LAMEIRAO":"Lameirão",
    "VALLADAO":"Valladão","SODRE":"Sodré","JANDRE":"Jandré","FELICIO":"Felício","BOECHAT":"Boechat",
    "DEMAUIR":"Demauir","ICARAI":"Icaraí","NITEROI":"Niterói","TERESOPOLIS":"Teresópolis",
    "HUMAITA":"Humaitá","MACAE":"Macaé","VITORIA":"Vitória","GLORIA":"Glória","GONCALO":"Gonçalo",
    "FATIMA":"Fátima","IGUACU":"Iguaçu","GALEAO":"Galeão","SAUDE":"Saúde","CLINICA":"Clínica",
    "CLINICAS":"Clínicas","CANCER":"Câncer","FORCA":"Força","AEREA":"Aérea",
    "PEDIATRICO":"Pediátrico","CARDIACO":"Cardíaco","CIRURGICO":"Cirúrgico",
    "NEUROLOGICO":"Neurológico","PUBLICO":"Público","LUCIA":"Lúcia","VANIA":"Vânia",
    "THAINA":"Thaína","GREGORIO":"Gregório","BERTELI":"Berteli","JULIANO":"Juliano",
    "EDSON":"Edson","VARGAS":"Vargas","NETTO":"Netto","FIGUEIREDO":"Figueiredo",
}
STOP = {"de","da","do","das","dos","e","a","o","em","no","na","nos","nas","ao","aos"}

def normalize_name(s):
    if not s: return s
    parts = re.split(r"(\s+|[-/])", s.strip())
    out = []
    for i, p in enumerate(parts):
        if p.strip() == "": out.append(p); continue
        if p in ("-", "/"): out.append(f" {p} "); continue
        w = p.strip(); up = w.upper(); fixed = WORD_FIX.get(up)
        if fixed: out.append(fixed)
        else:
            cap = w.capitalize()
            out.append(cap.lower() if cap.lower() in STOP and i > 0 else cap)
    return "".join(out).strip()

def normalize_hospital(s):
    if not s: return s
    key = re.sub(r"\s+", " ", s.strip().upper())
    if key in HOSPITAL_MAP: return HOSPITAL_MAP[key]
    return normalize_name(s)

FIELD_ALIASES = {
    "empresa":["EMPRESA"],"data":["DATA","DATA CIRURGIA","DATA_CIRURGIA"],
    "paciente":["PACIENTE","NOME_PACIENTE","NOME DO PACIENTE"],"hospital":["HOSPITAL"],
    "medico":["MÉDICO CORRETO","MEDICO CORRETO","MÉDICO","MEDICO"],
    "valor":["TOTAL","Total Geral verificado","Total Geral"],
    "classificacao":["CLASSIFICAÇÃO","CLASSIFICACAO","CLASSIFICAÇAO"],
}
CRANIO_KEYWORDS = ["neuro","crânio","cranio"]
COLUNA_KEYWORDS = ["coluna"]
EMERG_DISCOUNT = 0.10
EMERG_TOKENS = {"EMERGENCIA","EMERGÊNCIA","URGENCIA","URGÊNCIA"}
DOCTOR_ALIAS = {
    "GUSTAVO DA FONTOURA GALVAO":"Gustavo da Fontoura Galvão",
    "ROGERIO MARTINS PIRES DE AMORIM":"Rogério Amorim","ROGERIO AMORIM":"Rogério Amorim",
}

# === REGRA L.COSTA (3S-001) ===
TAXA_DEFAULT = 0.20
TAXA_OVERRIDE = {
    ("JULIANO BERTELI DE FIGUEIREDO",  "coluna"): 0.25,
    ("EDSON VARGAS DE OLIVEIRA NETTO", "coluna"): 0.25,
}

def normalize_key(s): return re.sub(r"\s+"," ", str(s or "").strip().upper())

def parse_date(val):
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    if val is None: return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%d/%m/%y"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def fmt_date(d): return d.strftime("%d/%m/%Y") if d else ""

def detect_header(ws):
    for row_idx in range(1, 11):
        row = [str(ws.cell(row_idx,c).value or "").strip() for c in range(1, ws.max_column+1)]
        col_map = {}
        for field, aliases in FIELD_ALIASES.items():
            for a in aliases:
                for ci, cell in enumerate(row):
                    if normalize_key(cell) == normalize_key(a):
                        col_map[field] = ci+1; break
                if field in col_map: break
        if len(col_map) >= 3: return row_idx, col_map
    return None, {}

def _norm_token(s):
    import unicodedata
    no_acc = "".join(ch for ch in unicodedata.normalize("NFKD", str(s or "")) if not unicodedata.combining(ch))
    return re.sub(r"\s+"," ", no_acc).upper().strip()

def read_sheet(ws):
    header_row, col_map = detect_header(ws)
    if not header_row: return []
    rows = []; emerg_count = 0
    for r in range(header_row+1, ws.max_row+1):
        def g(f): return ws.cell(r, col_map[f]).value if f in col_map else None
        paciente = str(g("paciente") or "").strip()
        medico = str(g("medico") or "").strip()
        if medico: medico = DOCTOR_ALIAS.get(_norm_token(medico), medico)
        hospital = str(g("hospital") or "").strip()
        empresa = str(g("empresa") or "L.COSTA").strip()
        classif = str(g("classificacao") or "").strip()
        raw_val = g("valor"); data = parse_date(g("data"))
        if isinstance(raw_val, (int,float)): valor = float(raw_val)
        else:
            try: valor = float(str(raw_val).replace("R$","").replace(".","").replace(",",".").strip())
            except: valor = 0.0
        if not paciente or valor == 0.0: continue
        is_emerg = _norm_token(classif) in {_norm_token(t) for t in EMERG_TOKENS}
        if is_emerg:
            valor = round(valor * (1 - EMERG_DISCOUNT), 2); emerg_count += 1
        rows.append(dict(paciente=paciente, medico=medico, hospital=hospital,
                         empresa=empresa, data=data, valor=valor,
                         classificacao=classif, emergencia=is_emerg))
    if emerg_count:
        print(f"    [EMERGÊNCIA] {emerg_count} cirurgias com -10% em '{ws.title}'")
    return rows

def merge_rows(all_rows, especialidade):
    bucket = {}
    for r in all_rows:
        key = normalize_key(r["paciente"])
        if key not in bucket:
            bucket[key] = dict(paciente=r["paciente"], hospital=r["hospital"],
                               empresa=r["empresa"], data=r["data"],
                               valor=0.0, especialidade=especialidade)
        bucket[key]["valor"] += r["valor"]
        if r["data"] and (bucket[key]["data"] is None or r["data"] < bucket[key]["data"]):
            bucket[key]["data"] = r["data"]
    return sorted(bucket.values(), key=lambda x: x["data"] or date(1900,1,1))

def get_taxa_rate(medico_name, especialidade):
    return TAXA_OVERRIDE.get((_norm_token(medico_name), especialidade), TAXA_DEFAULT)

def consolidate_by_patient(cranio_rows, coluna_rows, medico_name):
    """Une linhas do MESMO paciente (mesmo medico) — incluindo cruzando cranio + coluna.
    Taxa eh calculada por linha com a aliquota da especialidade ANTES da soma, pra preservar
    a regra de override (ex: Juliano Coluna 25% nao se mistura com 20% de cranio)."""
    bucket = {}
    for esp, rows in (("cranio", cranio_rows), ("coluna", coluna_rows)):
        rate = get_taxa_rate(medico_name, esp)
        for r in rows:
            key = normalize_key(r["paciente"])
            if key not in bucket:
                bucket[key] = dict(paciente=r["paciente"], hospital=r["hospital"],
                                   empresa=r["empresa"], data=r["data"],
                                   valor=0.0, taxa=0.0)
            bucket[key]["valor"] += r["valor"]
            bucket[key]["taxa"]  += round(r["valor"] * rate, 2)
            if r["data"] and (bucket[key]["data"] is None or r["data"] < bucket[key]["data"]):
                bucket[key]["data"] = r["data"]
    return sorted(bucket.values(), key=lambda x: x["data"] or date(1900, 1, 1))

def doctor_has_override(medico_name):
    nm = _norm_token(medico_name)
    return any(k[0] == nm for k in TAXA_OVERRIDE.keys())

COLS = [("MÉDICO",30),("L.COSTA",10),("DATA",13),("NOME DO PACIENTE",36),
        ("HOSPITAL",30),("VALOR TOTAL",16),("TAXA (20%)",14)]
VAL_COLS = {6, 7}

def setup_columns(ws, empresa):
    cols = list(COLS); cols[1] = (empresa, COLS[1][1])
    for ci, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def build_doctor_block(ws, start_row, medico_name, cranio_rows, coluna_rows, empresa):
    patients = consolidate_by_patient(cranio_rows, coluna_rows, medico_name)
    if not patients: return start_row, 0.0, 0.0

    has_override = doctor_has_override(medico_name)
    taxa_col_header = "TAXA" if has_override else "TAXA (20%)"

    cols_local = list(COLS)
    cols_local[1] = (empresa, COLS[1][1])
    cols_local[6] = (taxa_col_header, COLS[6][1])

    bdr = border_all(C["border"]); bdr_hdr = border_all(C["border_hdr"])

    r = start_row
    ws.row_dimensions[r].height = 24
    for ci, (hdr, _) in enumerate(cols_local, 1):
        c = ws.cell(r, ci, hdr)
        c.fill = fill(C["header_bg"]); c.font = fnt(C["txt_head"], bold=True, size=12)
        c.alignment = aln("center"); c.border = bdr_hdr

    total_valor = 0.0; total_taxa = 0.0
    for p in patients:
        r += 1
        ws.row_dimensions[r].height = 20
        taxa_row = p["taxa"]   # ja consolidada por especialidade em consolidate_by_patient
        total_valor += p["valor"]; total_taxa += taxa_row
        vals = [normalize_name(medico_name), empresa, fmt_date(p["data"]),
                normalize_name(p["paciente"]), normalize_hospital(p["hospital"]),
                p["valor"], taxa_row]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.fill = fill(C["val_bg"] if ci in VAL_COLS else C["row_bg"])
            c.font = fnt(C["txt_dark"], size=12); c.border = bdr
            if ci in VAL_COLS:
                c.alignment = aln("right"); c.number_format = '#,##0.00'
                if ci == 7: c.font = fnt(C["txt_gold"], bold=True, size=12)
            elif ci in (2, 3): c.alignment = aln("center")
            else: c.alignment = aln("left")

    r += 1
    ws.row_dimensions[r].height = 4
    for ci in range(1, len(COLS)+1):
        c = ws.cell(r, ci, "")
        c.fill = fill(C["sep_bg"])
        c.border = Border(top=Side(style="medium", color=C["border_hdr"]))

    r += 1
    ws.row_dimensions[r].height = 22
    for ci in range(1, len(COLS)+1):
        c = ws.cell(r, ci, ""); c.fill = fill(C["total_bg"]); c.border = bdr_hdr
    ws.cell(r, 5, "TOTAL").font = fnt(C["txt_dim"], bold=True, size=12)
    ws.cell(r, 5).alignment = aln("right"); ws.cell(r, 5).fill = fill(C["total_bg"])
    tc = ws.cell(r, 6, total_valor)
    tc.fill = fill(C["val_bg"]); tc.font = fnt(C["txt_dark"], bold=True, size=13)
    tc.alignment = aln("right"); tc.number_format = '#,##0.00'; tc.border = bdr_hdr
    tc_taxa = ws.cell(r, 7, total_taxa)
    tc_taxa.fill = fill(C["val_bg"]); tc_taxa.font = fnt(C["txt_gold"], bold=True, size=13)
    tc_taxa.alignment = aln("right"); tc_taxa.number_format = '#,##0.00'; tc_taxa.border = bdr_hdr

    return r + 1, total_valor, total_taxa

def detect_tabs(wb):
    names = wb.sheetnames
    cranio = next((n for n in reversed(names) if any(k in n.lower() for k in CRANIO_KEYWORDS) and 'v1' not in n.lower()), None)
    coluna = next((n for n in reversed(names) if any(k in n.lower() for k in COLUNA_KEYWORDS) and 'v1' not in n.lower()), None)
    return cranio, coluna

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True,
                    help="Caminho completo do xlsx de saída (ex: .../Tirinhas_LCOSTA_Abril_2026.xlsx)")
    ap.add_argument("--empresa", default="L.COSTA")
    args = ap.parse_args()

    wb_src = openpyxl.load_workbook(args.input, data_only=True)
    cranio_tab, coluna_tab = detect_tabs(wb_src)
    print(f"Crânio  -> {cranio_tab}")
    print(f"Coluna  -> {coluna_tab}")

    cranio_all = read_sheet(wb_src[cranio_tab]) if cranio_tab and cranio_tab in wb_src.sheetnames else []
    coluna_all = read_sheet(wb_src[coluna_tab]) if coluna_tab and coluna_tab in wb_src.sheetnames else []

    by_medico = defaultdict(lambda: {"cranio": [], "coluna": []})
    for r in cranio_all:
        if r["medico"]: by_medico[normalize_key(r["medico"])]["cranio"].append(r)
    for r in coluna_all:
        if r["medico"]: by_medico[normalize_key(r["medico"])]["coluna"].append(r)

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Tirinhas"
    setup_columns(ws, args.empresa)

    print(f"\nGerando {len(by_medico)} blocos em: {args.output}\n")

    def canonical_name(rows_c, rows_col):
        rows = rows_c or rows_col
        return rows[0]["medico"] if rows else "Desconhecido"

    next_row = 1
    total_geral = 0.0; total_taxa_geral = 0.0
    for norm_name, data in sorted(by_medico.items()):
        name = canonical_name(data["cranio"], data["coluna"])
        next_row, valor, taxa = build_doctor_block(
            ws, next_row, name, data["cranio"], data["coluna"], args.empresa
        )
        next_row += 1
        total_geral += valor; total_taxa_geral += taxa
        cr = len(data["cranio"]); co = len(data["coluna"])
        tag = (f"Cr:{cr} " if cr else "") + (f"Co:{co}" if co else "")
        flag = " *" if doctor_has_override(name) else ""
        print(f"  {normalize_name(name):<48} {tag:<10} R$ {valor:>12,.2f}  Taxa R$ {taxa:>10,.2f}{flag}")

    out_dir = os.path.dirname(args.output)
    if out_dir: os.makedirs(out_dir, exist_ok=True)
    wb_out.save(args.output)

    print(f"\n{'─'*88}")
    print(f"  Empresa    : {args.empresa}")
    print(f"  Médicos    : {len(by_medico)}")
    print(f"  Volume     : R$ {total_geral:>12,.2f}")
    print(f"  Taxa total : R$ {total_taxa_geral:>12,.2f}")
    print(f"  Arquivo    : {args.output}")
    print(f"  '*' = medico com override de taxa especifico")

if __name__ == "__main__":
    main()
