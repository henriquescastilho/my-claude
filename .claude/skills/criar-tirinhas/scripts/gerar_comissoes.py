#!/usr/bin/env python3
"""
Gerar Comissões dos Agentes — L.COSTA (3S-001)

Saída: 1 xlsx único consolidado com uma aba por agente (+ aba "SEM AGENTE" se houver).

Regras (L.COSTA):
- 4% sobre o LÍQUIDO (TOTAL GERAL - DESCONTO) — base bruta da cirurgia.
- 3,5% em cirurgias de COLUNA do Juliano Berteli de Figueiredo.
- 3,5% em cirurgias de COLUNA do Edson Vargas de Oliveira Netto.
- Bônus 0,25% para Luiz Henrique Castilho Viana de Castilho, calculado sobre a soma
  do TOTAL GERAL - DESCONTO das DUAS empresas (L.COSTA + CARE SA) no mês.
  Pago pela L.COSTA.
  Para calcular o pedaço da CARE SA, passar --caresa-input <caminho_planilha_caresa>.
- Cirurgias sem AGENTE preenchido vão para a aba "SEM AGENTE" (sem cálculo).
- Coluna `Bônus (0,25%)` aparece em todas as abas mas só preenchida na aba do Luiz.
"""

import argparse, os, re, unicodedata
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = {
    "header_bg": "C8E6C9", "row_bg": "F1F8E9", "val_bg": "FFFFFF",
    "total_bg":  "DCEDC8", "sep_bg": "E8F5E9",
    "txt_dark":  "1A1A1A", "txt_head":"1B5E20", "txt_dim":"33691E",
    "txt_com":   "2E7D32", "txt_bonus":"F57F17",
    "border":    "BDBDBD", "border_hdr":"9E9E9E",
}

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def fnt(hex_, bold=False, size=12): return Font(color=hex_, bold=bold, size=size, name="Calibri")
def aln(h="left", v="center"): return Alignment(horizontal=h, vertical=v)
def border_all(color="BDBDBD"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _norm_token(s):
    no_acc = "".join(ch for ch in unicodedata.normalize("NFKD", str(s or ""))
                     if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", no_acc).upper().strip()

def normalize_key(s):
    return re.sub(r"\s+", " ", str(s or "").strip().upper())

def parse_date(val):
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    if val is None: return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def fmt_date(d):
    return d.strftime("%d/%m/%Y") if d else ""

FIELD_ALIASES = {
    "data":          ["DATA", "DATA CIRURGIA", "DATA_CIRURGIA"],
    "paciente":      ["PACIENTE", "NOME_PACIENTE", "NOME DO PACIENTE"],
    "hospital":      ["HOSPITAL"],
    "medico":        ["MÉDICO CORRETO", "MEDICO CORRETO", "MÉDICO", "MEDICO"],
    "valor_liquido": ["TOTAL GERAL - DESCONTO", "TOTAL GERAL-DESCONTO"],
    "valor_cheio":   ["TOTAL", "Total Geral verificado", "Total Geral"],
    "agente":        ["AGENTE CORRETO", "AGENTE", "AGENTE_CORRETO"],
    "classificacao": ["CLASSIFICAÇÃO", "CLASSIFICACAO", "CLASSIFICAÇAO"],
}

CRANIO_KEYWORDS = ["neuro", "crânio", "cranio"]
COLUNA_KEYWORDS = ["coluna"]

# === REGRA L.COSTA ===
COMMISSION_DEFAULT = 0.04

# Override por (medico_normalizado, especialidade) -> rate
COMMISSION_OVERRIDE_BY_DOC_ESP = {
    ("JULIANO BERTELI DE FIGUEIREDO",  "coluna"): 0.035,
    ("EDSON VARGAS DE OLIVEIRA NETTO", "coluna"): 0.035,
}

# Bônus 0,25% sobre soma das duas empresas, pago ao Luiz Henrique pela L.COSTA.
BONUS_RATE   = 0.0025
BONUS_TARGET = "LUIZ HENRIQUE CASTILHO VIANA DE CASTILHO"

# Despesa fixa mensal por agente (pago todo mês, independente de cirurgia).
# L.COSTA paga 2 agentes: Luiz Henrique + Saint Clair.
DESPESA_FIXA_POR_AGENTE = 2200.00

EMERG_TOKENS = {"EMERGENCIA", "EMERGÊNCIA", "URGENCIA", "URGÊNCIA"}

DOCTOR_ALIAS = {
    "GUSTAVO DA FONTOURA GALVAO":              "Gustavo da Fontoura Galvão",
    "ROGERIO MARTINS PIRES DE AMORIM":         "Rogério Amorim",
    "ROGERIO AMORIM":                          "Rogério Amorim",
}

VENDORS = {
    "SAINT CLAIR BARBOSA DE OLIVEIRA":            "Saint Clair Barbosa de Oliveira",
    # Luiz Henrique aparece em DUAS grafias na planilha SPI:
    #   - "Luiz Henrique Castilho Viana de Castilho" (forma contrato/canonical)
    #   - "Luiz Henrique Viana de Castilho"          (forma curta da SPI — falta "Castilho" do meio)
    # As duas mapeiam pro mesmo agente canonical.
    "LUIZ HENRIQUE CASTILHO VIANA DE CASTILHO":   "Luiz Henrique Castilho Viana de Castilho",
    "LUIZ HENRIQUE VIANA DE CASTILHO":            "Luiz Henrique Castilho Viana de Castilho",
}

def parse_vendor(agente_raw):
    if not agente_raw: return None
    s = re.sub(r"^\s*CONTRACTOR\s*[-–]\s*", "", str(agente_raw), flags=re.IGNORECASE).strip()
    return VENDORS.get(_norm_token(s))

def get_commission_rate(medico_name, especialidade):
    return COMMISSION_OVERRIDE_BY_DOC_ESP.get(
        (_norm_token(medico_name), especialidade),
        COMMISSION_DEFAULT,
    )

def normalize_name_simple(s):
    if not s: return s
    parts = re.split(r"(\s+|[-/])", str(s).strip())
    stop = {"de","da","do","das","dos","e","a","o","em","no","na","ao","aos"}
    out = []
    for i, p in enumerate(parts):
        if p.strip() == "" or p in ("-","/"):
            out.append(p); continue
        low = p.lower()
        if i > 0 and low in stop:
            out.append(low)
        else:
            out.append(p.capitalize() if (p.isupper() or p.islower()) else p)
    return "".join(out).strip()

def detect_header(ws):
    for row_idx in range(1, 11):
        row = [str(ws.cell(row_idx, c).value or "").strip() for c in range(1, ws.max_column+1)]
        col_map = {}
        for field, aliases in FIELD_ALIASES.items():
            for a in aliases:
                for ci, cell in enumerate(row):
                    if normalize_key(cell) == normalize_key(a):
                        col_map[field] = ci + 1
                        break
                if field in col_map: break
        if len(col_map) >= 3:
            return row_idx, col_map
    return None, {}

def detect_tabs(wb):
    names = wb.sheetnames
    def pick(keywords):
        cand_no_v1 = [n for n in reversed(names)
                      if any(k in n.lower() for k in keywords) and 'v1' not in n.lower()]
        if cand_no_v1: return cand_no_v1[0]
        cand_any = [n for n in reversed(names)
                    if any(k in n.lower() for k in keywords)]
        return cand_any[0] if cand_any else None
    return pick(CRANIO_KEYWORDS), pick(COLUNA_KEYWORDS)

def read_sheet(ws, especialidade):
    header_row, col_map = detect_header(ws)
    if not header_row: return []
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        def g(f): return ws.cell(r, col_map[f]).value if f in col_map else None
        paciente = str(g("paciente") or "").strip()
        medico   = str(g("medico")   or "").strip()
        if medico:
            medico = DOCTOR_ALIAS.get(_norm_token(medico), medico)
        hospital = str(g("hospital") or "").strip()
        agente   = str(g("agente")   or "").strip()
        classif  = str(g("classificacao") or "").strip()
        raw_val  = g("valor_cheio")
        if raw_val is None:
            raw_val = g("valor_liquido")
        data     = parse_date(g("data"))
        if isinstance(raw_val, (int, float)):
            valor = float(raw_val)
        else:
            try:
                valor = float(str(raw_val).replace("R$","").replace(".","").replace(",",".").strip())
            except:
                valor = 0.0
        if not paciente or valor == 0.0: continue
        is_emerg = _norm_token(classif) in {_norm_token(t) for t in EMERG_TOKENS}
        rows.append(dict(
            paciente=paciente, medico=medico, hospital=hospital,
            data=data, valor=valor, agente=agente,
            vendedor=parse_vendor(agente), especialidade=especialidade,
            classificacao=classif, emergencia=is_emerg,
        ))
    return rows

def soma_liquido(input_path):
    """Lê uma planilha SPI e retorna a soma de TOTAL GERAL - DESCONTO das abas operacionais."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    cranio_tab, coluna_tab = detect_tabs(wb)
    cranio = read_sheet(wb[cranio_tab], "cranio") if cranio_tab else []
    coluna = read_sheet(wb[coluna_tab], "coluna") if coluna_tab else []
    return sum(r["valor"] for r in cranio + coluna)

COLS = [
    ("AGENTE",            30),
    ("DATA",              13),
    ("MÉDICO",            32),
    ("NOME DO PACIENTE",  36),
    ("HOSPITAL",          30),
    ("ESPECIALIDADE",     14),
    ("VALOR (TOTAL)",     16),
    ("COMISSÃO",          16),
    ("BÔNUS (0,25%)",     16),
]
VAL_COLS = {7, 8, 9}

def build_sheet(ws, vendor_display, rows, with_commission=True, bonus_value=None):
    ws.title = vendor_display[:31]

    for ci, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    bdr = border_all(C["border"])
    bdr_hdr = border_all(C["border_hdr"])

    r = 1
    ws.row_dimensions[r].height = 24
    for ci, (hdr, _) in enumerate(COLS, 1):
        c = ws.cell(r, ci, hdr)
        c.fill = fill(C["header_bg"])
        c.font = fnt(C["txt_head"], bold=True, size=12)
        c.alignment = aln("center")
        c.border = bdr_hdr

    rows_sorted = sorted(rows, key=lambda x: (
        0 if x["especialidade"] == "cranio" else 1,
        x["data"] or date(1900, 1, 1),
    ))
    total_valor = 0.0
    total_com   = 0.0
    for p in rows_sorted:
        r += 1
        ws.row_dimensions[r].height = 20
        if with_commission:
            rate = get_commission_rate(p["medico"], p["especialidade"])
            com  = round(p["valor"] * rate, 2)
        else:
            com  = ""
        total_valor += p["valor"]
        if isinstance(com, (int, float)):
            total_com += com

        vals = [
            vendor_display,
            fmt_date(p["data"]),
            normalize_name_simple(p["medico"]),
            normalize_name_simple(p["paciente"]),
            normalize_name_simple(p["hospital"]),
            "Crânio" if p["especialidade"] == "cranio" else "Coluna",
            p["valor"],
            com,
            "",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci, v)
            c.fill = fill(C["val_bg"] if ci in VAL_COLS else C["row_bg"])
            c.font = fnt(C["txt_dark"], size=12)
            c.border = bdr
            if ci in VAL_COLS:
                c.alignment = aln("right")
                c.number_format = '#,##0.00'
                if ci == 8:
                    c.font = fnt(C["txt_com"], bold=True, size=12)
                elif ci == 9:
                    c.font = fnt(C["txt_bonus"], bold=True, size=12)
            elif ci in (2, 6):
                c.alignment = aln("center")
            else:
                c.alignment = aln("left")

    r += 1
    ws.row_dimensions[r].height = 4
    for ci in range(1, len(COLS)+1):
        c = ws.cell(r, ci, "")
        c.fill = fill(C["sep_bg"])
        c.border = Border(top=Side(style="medium", color=C["border_hdr"]))

    r += 1
    ws.row_dimensions[r].height = 22
    for ci in range(1, len(COLS)+1):
        c = ws.cell(r, ci, "")
        c.fill = fill(C["total_bg"])
        c.border = bdr_hdr
    ws.cell(r, 6, "TOTAL").font = fnt(C["txt_dim"], bold=True, size=12)
    ws.cell(r, 6).alignment = aln("right")
    tc = ws.cell(r, 7, total_valor)
    tc.font = fnt(C["txt_dark"], bold=True, size=13)
    tc.alignment = aln("right")
    tc.number_format = '#,##0.00'
    if with_commission:
        cc = ws.cell(r, 8, total_com)
        cc.font = fnt(C["txt_com"], bold=True, size=13)
        cc.alignment = aln("right")
        cc.number_format = '#,##0.00'
    if bonus_value is not None:
        bc = ws.cell(r, 9, bonus_value)
        bc.font = fnt(C["txt_bonus"], bold=True, size=13)
        bc.alignment = aln("right")
        bc.number_format = '#,##0.00'

    return total_valor, total_com

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",        required=True, help="Planilha LCOSTA - MM.YYYY N.xlsx")
    ap.add_argument("--output",       required=True, help="Caminho do xlsx de saída (Comissoes_<Mes>_<Ano>.xlsx)")
    ap.add_argument("--caresa-input", default=None,  help="(Opcional) Planilha CARE SA - MM.YYYY N.xlsx para cálculo completo do bônus 0,25%%")
    args = ap.parse_args()

    wb_src = openpyxl.load_workbook(args.input, data_only=True)
    cranio_tab, coluna_tab = detect_tabs(wb_src)
    print(f"Crânio  -> {cranio_tab}")
    print(f"Coluna  -> {coluna_tab}")

    cranio = read_sheet(wb_src[cranio_tab], "cranio") if cranio_tab else []
    coluna = read_sheet(wb_src[coluna_tab], "coluna") if coluna_tab else []
    all_rows = cranio + coluna

    by_vendor = defaultdict(list)
    sem_agente = []
    for row in all_rows:
        if row["vendedor"]:
            by_vendor[row["vendedor"]].append(row)
        else:
            sem_agente.append(row)

    soma_lcosta = sum(r["valor"] for r in all_rows)
    if args.caresa_input:
        try:
            soma_caresa = soma_liquido(args.caresa_input)
            print(f"\n[BÔNUS 0,25%]")
            print(f"  Soma líquido L.COSTA : R$ {soma_lcosta:>13,.2f}")
            print(f"  Soma líquido CARE SA : R$ {soma_caresa:>13,.2f}")
            base_bonus = soma_lcosta + soma_caresa
        except Exception as e:
            print(f"[AVISO] Não foi possível ler --caresa-input ({e}). Bônus calculado APENAS sobre L.COSTA — completar manualmente.")
            base_bonus = soma_lcosta
    else:
        print(f"\n[BÔNUS 0,25%] --caresa-input não passado.")
        print(f"  Bônus calculado APENAS sobre L.COSTA (R$ {soma_lcosta:,.2f}). Completar manualmente com a parcela CARE SA.")
        base_bonus = soma_lcosta

    bonus_luiz = round(base_bonus * BONUS_RATE, 2)
    print(f"  Base bônus           : R$ {base_bonus:>13,.2f}")
    print(f"  Bônus Luiz (0,25%)   : R$ {bonus_luiz:>13,.2f}")

    luiz_canonical = VENDORS.get(BONUS_TARGET) or "Luiz Henrique Castilho Viana de Castilho"

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    print(f"\nGerando comissões consolidadas...\n")
    total_geral_val = 0.0
    total_geral_com = 0.0

    for vendor in sorted(by_vendor.keys()):
        rows = by_vendor[vendor]
        ws = wb_out.create_sheet()
        bonus_param = bonus_luiz if _norm_token(vendor) == _norm_token(luiz_canonical) else None
        valor, com = build_sheet(ws, vendor, rows, with_commission=True, bonus_value=bonus_param)
        total_geral_val += valor
        total_geral_com += com
        bonus_str = f"  Bonus R$ {bonus_param:>10,.2f}" if bonus_param is not None else ""
        print(f"  {vendor:<45} {len(rows):3d} cir.  Vol R$ {valor:>13,.2f}  Com R$ {com:>10,.2f}{bonus_str}")

    if sem_agente:
        ws = wb_out.create_sheet()
        valor, _ = build_sheet(ws, "SEM AGENTE", sem_agente, with_commission=False)
        print(f"  {'SEM AGENTE':<45} {len(sem_agente):3d} cir.  Vol R$ {valor:>13,.2f}  (sem comissão)")
    else:
        print(f"  (nenhuma cirurgia sem AGENTE)")

    out_dir = os.path.dirname(args.output)
    if out_dir: os.makedirs(out_dir, exist_ok=True)
    wb_out.save(args.output)

    print(f"\n{'─'*78}")
    print(f"  Empresa           : L.COSTA (3S-001)")
    print(f"  Agentes           : {len(by_vendor)}")
    print(f"  Volume total      : R$ {total_geral_val:>13,.2f}")
    print(f"  Comissão total    : R$ {total_geral_com:>13,.2f}")
    print(f"  Bônus Luiz 0,25%  : R$ {bonus_luiz:>13,.2f}")
    print(f"  Custo total       : R$ {total_geral_com + bonus_luiz:>13,.2f}")
    print(f"  Sem agente        : {len(sem_agente)} cirurgias")
    print(f"  Arquivo           : {args.output}")

if __name__ == "__main__":
    main()
