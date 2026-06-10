#!/usr/bin/env python3
"""
Gera Tirinhas em PDF — CARE SA — paisagem A4.
Mesma regra de Taxa do gerar_tirinhas.py (xlsx) com overrides Gustavo Adolpho
(20,88% Cr / 25,05% Co) e Silas Augusto (15% Cr / 15% Co, com -10% silencioso permanente no cranio).
Quebras de página limpas: bloco tenta caber em 1 página (KeepTogether), e se
maior que uma página o header repete no topo (repeatRows=1).
"""

import argparse, os, re
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, KeepTogether
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

HOSPITAL_MAP = {
    "CASA DE PORTUGAL":"Casa de Portugal","CASA SAUDE S JOSE HUMAITA":"Casa de Saúde São José Humaitá",
    "CLINICA SAO GONCALO":"Clínica São Gonçalo","CLINICA SAO VICENTE":"Clínica São Vicente",
    "GLORIA D OR":"Glória D'Or","GLORIA DOR":"Glória D'Or","HOSP BADIM":"Hospital Badim",
    "HOSP BANGU":"Hospital Bangu","HOSP BARRA DOR II NOVO":"Hospital Barra D'Or II Novo",
    "HOSP CAXIAS DOR":"Hospital Caxias D'Or","HOSP CHN":"Hospital CHN",
    "HOSP COPA D OR":"Hospital Copa D'Or","HOSP COPA STAR":"Hospital Copa Star",
    "HOSP DE FORCA AEREA DO GALEAO":"Hospital de Força Aérea do Galeão",
    "HOSP NORTE DOR CASCADURA":"Hospital Norte D'Or Cascadura","HOSP OESTE DOR":"Hospital Oeste D'Or",
    "HOSP PASTEUR":"Hospital Pasteur","HOSP PRO CARDIACO":"Hospital Pró-Cardíaco",
    "HOSP QUINTA DOR":"Hospital Quinta D'Or","HOSP RIOS DOR":"Hospital Rios D'Or",
    "HOSP SAMARITANO":"Hospital Samaritano","HOSP SAMARITANO - BOTAFOGO":"Hospital Samaritano - Botafogo",
    "HOSP SAO LUCAS-COPACABANA":"Hospital São Lucas - Copacabana",
    "HOSP UNIMED ARARUAMA":"Hospital Unimed Araruama","HOSP UNIMED BARRA":"Hospital Unimed Barra",
    "HOSP VITORIA RJ":"Hospital Vitória - RJ","NITEROI D OR":"Niterói D'Or","PRONTONIL":"Prontonil",
}
WORD_FIX = {
    "HOSP":"Hospital","MAT":"Maternidade","SAO":"São","STA":"Santa","STO":"Santo","SRA":"Senhora",
    "NSR":"Nossa Senhora","ADAO":"Adão","ANTONIO":"Antônio","JOAO":"João","JOSE":"José",
    "JULIO":"Júlio","MARIO":"Mário","OTAVIO":"Otávio","OCTAVIO":"Otávio","ROGERIO":"Rogério",
    "SAVIO":"Sávio","VINICIUS":"Vinícius","CLAUDIO":"Cláudio","MARCIO":"Márcio","FLAVIO":"Flávio",
    "FABIO":"Fábio","LUCIO":"Lúcio","CAIO":"Caio","AURELIO":"Aurélio","HELIO":"Hélio",
    "SERGIO":"Sérgio","CICERO":"Cícero","HENRIQUE":"Henrique","CORREA":"Corrêa","SIMOES":"Simões",
    "GUIMARAES":"Guimarães","GONCALVES":"Gonçalves","MOURAO":"Mourão","LAMEIRAO":"Lameirão",
    "VALLADAO":"Valladão","SODRE":"Sodré","JANDRE":"Jandré","FELICIO":"Felício","BOECHAT":"Boechat",
    "DEMAUIR":"Demauir","ICARAI":"Icaraí","NITEROI":"Niterói","TERESOPOLIS":"Teresópolis",
    "HUMAITA":"Humaitá","MACAE":"Macaé","VITORIA":"Vitória","GLORIA":"Glória","GONCALO":"Gonçalo",
    "FATIMA":"Fátima","IGUACU":"Iguaçu","GALEAO":"Galeão","SAUDE":"Saúde","CLINICA":"Clínica",
    "CLINICAS":"Clínicas","CANCER":"Câncer","FORCA":"Força","AEREA":"Aérea",
    "PEDIATRICO":"Pediátrico","CARDIACO":"Cardíaco","CIRURGICO":"Cirúrgico",
    "NEUROLOGICO":"Neurológico","PUBLICO":"Público","LUCIA":"Lúcia","VANIA":"Vânia",
    "THAINA":"Thaína","GREGORIO":"Gregório","ADOLPHO":"Adolpho","CANNABRAVA":"Cannabrava",
}
STOP = {"de","da","do","das","dos","e","a","o","em","no","na","nos","nas","ao","aos"}

def normalize_name(s):
    if not s: return s
    parts = re.split(r"(\s+|[-/])", s.strip())
    out = []
    for i, p in enumerate(parts):
        if p.strip() == "": out.append(p); continue
        if p in ("-", "/"): out.append(f" {p} "); continue
        w = p.strip(); up = w.upper(); fixed = WORD_FIX.get(up)
        if fixed: out.append(fixed)
        else:
            cap = w.capitalize()
            out.append(cap.lower() if cap.lower() in STOP and i > 0 else cap)
    return "".join(out).strip()

def normalize_hospital(s):
    if not s: return s
    key = re.sub(r"\s+", " ", s.strip().upper())
    if key in HOSPITAL_MAP: return HOSPITAL_MAP[key]
    return normalize_name(s)

FIELD_ALIASES = {
    "empresa":["EMPRESA"],"data":["DATA","DATA CIRURGIA","DATA_CIRURGIA"],
    "paciente":["PACIENTE","NOME_PACIENTE","NOME DO PACIENTE"],"hospital":["HOSPITAL"],
    "medico":["MÉDICO CORRETO","MEDICO CORRETO","MÉDICO","MEDICO"],
    "valor":["TOTAL","Total Geral verificado","Total Geral"],
    "classificacao":["CLASSIFICAÇÃO","CLASSIFICACAO","CLASSIFICAÇAO"],
}
CRANIO_KEYWORDS = ["neuro","crânio","cranio"]
COLUNA_KEYWORDS = ["coluna"]
EMERG_DISCOUNT = 0.10
EMERG_TOKENS = {"EMERGENCIA","EMERGÊNCIA","URGENCIA","URGÊNCIA"}
DOCTOR_ALIAS = {
    "GUSTAVO DA FONTOURA GALVAO":"Gustavo da Fontoura Galvão",
    "ROGERIO MARTINS PIRES DE AMORIM":"Rogério Amorim","ROGERIO AMORIM":"Rogério Amorim",
}

# === REGRA CARE SA ===
TAXA_DEFAULT = 0.20
TAXA_OVERRIDE = {
    ("GUSTAVO ADOLPHO CANNABRAVA CARVALHO","cranio"): 0.2088,
    ("GUSTAVO ADOLPHO CANNABRAVA CARVALHO","coluna"): 0.2505,
    ("SILAS AUGUSTO FARIA MARTINS",        "cranio"): 0.15,
    ("SILAS AUGUSTO FARIA MARTINS",        "coluna"): 0.15,
}

# (medico, especialidade) que SEMPRE leva -10% silencioso (mesmo em eletiva).
DESCONTO_SILENCIOSO_SEMPRE = {
    ("SILAS AUGUSTO FARIA MARTINS", "cranio"),
}

def normalize_key(s): return re.sub(r"\s+"," ", str(s or "").strip().upper())

def parse_date(val):
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    if val is None: return None
    s = str(val).strip()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%d/%m/%y"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def fmt_date(d): return d.strftime("%d/%m/%Y") if d else ""

def detect_header(ws):
    for row_idx in range(1, 11):
        row = [str(ws.cell(row_idx,c).value or "").strip() for c in range(1, ws.max_column+1)]
        col_map = {}
        for field, aliases in FIELD_ALIASES.items():
            for a in aliases:
                for ci, cell in enumerate(row):
                    if normalize_key(cell) == normalize_key(a):
                        col_map[field] = ci+1; break
                if field in col_map: break
        if len(col_map) >= 3: return row_idx, col_map
    return None, {}

def _norm_token(s):
    import unicodedata
    no_acc = "".join(ch for ch in unicodedata.normalize("NFKD", str(s or "")) if not unicodedata.combining(ch))
    return re.sub(r"\s+"," ", no_acc).upper().strip()

def read_sheet(ws, default_empresa, especialidade):
    header_row, col_map = detect_header(ws)
    if not header_row: return []
    rows = []
    for r in range(header_row+1, ws.max_row+1):
        def g(f): return ws.cell(r, col_map[f]).value if f in col_map else None
        paciente = str(g("paciente") or "").strip()
        medico = str(g("medico") or "").strip()
        if medico: medico = DOCTOR_ALIAS.get(_norm_token(medico), medico)
        hospital = str(g("hospital") or "").strip()
        empresa = str(g("empresa") or default_empresa).strip()
        classif = str(g("classificacao") or "").strip()
        raw_val = g("valor"); data = parse_date(g("data"))
        if isinstance(raw_val, (int,float)): valor = float(raw_val)
        else:
            try: valor = float(str(raw_val).replace("R$","").replace(".","").replace(",",".").strip())
            except: valor = 0.0
        if not paciente or valor == 0.0: continue
        is_emerg = _norm_token(classif) in {_norm_token(t) for t in EMERG_TOKENS}
        is_silencioso_sempre = (_norm_token(medico), especialidade) in DESCONTO_SILENCIOSO_SEMPRE
        if is_emerg or is_silencioso_sempre:
            valor = round(valor * (1 - EMERG_DISCOUNT), 2)
        rows.append(dict(paciente=paciente, medico=medico, hospital=hospital,
                         empresa=empresa, data=data, valor=valor,
                         classificacao=classif, emergencia=is_emerg))
    return rows

def merge_rows(all_rows, especialidade):
    bucket = {}
    for r in all_rows:
        key = normalize_key(r["paciente"])
        if key not in bucket:
            bucket[key] = dict(paciente=r["paciente"], hospital=r["hospital"],
                               empresa=r["empresa"], data=r["data"],
                               valor=0.0, especialidade=especialidade)
        bucket[key]["valor"] += r["valor"]
        if r["data"] and (bucket[key]["data"] is None or r["data"] < bucket[key]["data"]):
            bucket[key]["data"] = r["data"]
    return sorted(bucket.values(), key=lambda x: x["data"] or date(1900,1,1))

def get_taxa_rate(medico_name, especialidade):
    return TAXA_OVERRIDE.get((_norm_token(medico_name), especialidade), TAXA_DEFAULT)

def consolidate_by_patient(cranio_rows, coluna_rows, medico_name):
    """Une linhas do MESMO paciente (mesmo medico) — incluindo cruzando cranio + coluna.
    Taxa eh calculada por linha com a aliquota da especialidade ANTES da soma."""
    bucket = {}
    for esp, rows in (("cranio", cranio_rows), ("coluna", coluna_rows)):
        rate = get_taxa_rate(medico_name, esp)
        for r in rows:
            key = normalize_key(r["paciente"])
            if key not in bucket:
                bucket[key] = dict(paciente=r["paciente"], hospital=r["hospital"],
                                   empresa=r["empresa"], data=r["data"],
                                   valor=0.0, taxa=0.0)
            bucket[key]["valor"] += r["valor"]
            bucket[key]["taxa"]  += round(r["valor"] * rate, 2)
            if r["data"] and (bucket[key]["data"] is None or r["data"] < bucket[key]["data"]):
                bucket[key]["data"] = r["data"]
    return sorted(bucket.values(), key=lambda x: x["data"] or date(1900, 1, 1))

def doctor_has_override(medico_name):
    nm = _norm_token(medico_name)
    return any(k[0] == nm for k in TAXA_OVERRIDE.keys())

def detect_tabs(wb):
    names = wb.sheetnames
    cranio = next((n for n in reversed(names) if any(k in n.lower() for k in CRANIO_KEYWORDS) and 'v1' not in n.lower()), None)
    coluna = next((n for n in reversed(names) if any(k in n.lower() for k in COLUNA_KEYWORDS) and 'v1' not in n.lower()), None)
    return cranio, coluna

# ─── PDF ────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG = colors.HexColor("#FFE082")
COLOR_ROW_BG    = colors.HexColor("#FFFDE7")
COLOR_VAL_BG    = colors.HexColor("#FFFFFF")
COLOR_TOTAL_BG  = colors.HexColor("#FFF9C4")
COLOR_TXT_DARK  = colors.HexColor("#1A1A1A")
COLOR_TXT_HEAD  = colors.HexColor("#333333")
COLOR_TXT_DIM   = colors.HexColor("#666666")
COLOR_TXT_GOLD  = colors.HexColor("#8B6914")
COLOR_BORDER    = colors.HexColor("#BDBDBD")
COLOR_BORDER_HD = colors.HexColor("#9E9E9E")
COL_WIDTHS = [110, 50, 60, 165, 170, 90, 80]

_styles = getSampleStyleSheet()
STY_HEADER = ParagraphStyle('hdr', parent=_styles['Normal'], fontName='Helvetica-Bold',
                            fontSize=9, textColor=COLOR_TXT_HEAD, alignment=TA_CENTER, leading=11)
STY_CELL = ParagraphStyle('cell', parent=_styles['Normal'], fontName='Helvetica',
                          fontSize=9, textColor=COLOR_TXT_DARK, alignment=TA_LEFT, leading=11)
STY_CELL_C = ParagraphStyle('cellc', parent=STY_CELL, alignment=TA_CENTER)
STY_VAL = ParagraphStyle('val', parent=STY_CELL, alignment=TA_RIGHT)
STY_TAXA = ParagraphStyle('taxa', parent=STY_CELL, alignment=TA_RIGHT,
                          fontName='Helvetica-Bold', textColor=COLOR_TXT_GOLD)
STY_TOTAL_LBL = ParagraphStyle('totl', parent=STY_CELL, alignment=TA_RIGHT,
                               fontName='Helvetica-Bold', textColor=COLOR_TXT_DIM)
STY_TOTAL_VAL = ParagraphStyle('totv', parent=STY_CELL, alignment=TA_RIGHT,
                               fontName='Helvetica-Bold', fontSize=10, textColor=COLOR_TXT_DARK)
STY_TOTAL_TAX = ParagraphStyle('tott', parent=STY_CELL, alignment=TA_RIGHT,
                               fontName='Helvetica-Bold', fontSize=10, textColor=COLOR_TXT_GOLD)

def fmt_money(v):
    s = f"{v:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

def build_doctor_table(medico_name, cranio_rows, coluna_rows, empresa):
    patients = consolidate_by_patient(cranio_rows, coluna_rows, medico_name)
    if not patients: return None, 0.0, 0.0
    has_override = doctor_has_override(medico_name)
    taxa_col_header = "TAXA" if has_override else "TAXA (20%)"
    headers = ["MÉDICO", empresa, "DATA", "NOME DO PACIENTE", "HOSPITAL", "VALOR TOTAL", taxa_col_header]
    data = [[Paragraph(h, STY_HEADER) for h in headers]]
    total_valor = 0.0; total_taxa = 0.0
    for p in patients:
        taxa_row = p["taxa"]   # ja consolidada por especialidade
        total_valor += p["valor"]; total_taxa += taxa_row
        data.append([
            Paragraph(normalize_name(medico_name), STY_CELL),
            Paragraph(empresa, STY_CELL_C),
            Paragraph(fmt_date(p["data"]), STY_CELL_C),
            Paragraph(normalize_name(p["paciente"]), STY_CELL),
            Paragraph(normalize_hospital(p["hospital"]), STY_CELL),
            Paragraph(fmt_money(p["valor"]), STY_VAL),
            Paragraph(fmt_money(taxa_row), STY_TAXA),
        ])
    data.append([
        Paragraph("", STY_CELL), Paragraph("", STY_CELL), Paragraph("", STY_CELL), Paragraph("", STY_CELL),
        Paragraph("TOTAL", STY_TOTAL_LBL),
        Paragraph(fmt_money(total_valor), STY_TOTAL_VAL),
        Paragraph(fmt_money(total_taxa), STY_TOTAL_TAX),
    ])
    n_rows = len(data); last = n_rows - 1
    tbl = Table(data, colWidths=COL_WIDTHS, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), COLOR_HEADER_BG),
        ('GRID', (0,0), (-1,0), 0.7, COLOR_BORDER_HD),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,last-1), COLOR_ROW_BG),
        ('BACKGROUND', (5,1), (6,last-1), COLOR_VAL_BG),
        ('GRID', (0,1), (-1,last-1), 0.5, COLOR_BORDER),
        ('TOPPADDING', (0,1), (-1,last-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,last-1), 4),
        ('BACKGROUND', (0,last), (-1,last), COLOR_TOTAL_BG),
        ('BACKGROUND', (5,last), (6,last), COLOR_VAL_BG),
        ('GRID', (0,last), (-1,last), 0.7, COLOR_BORDER_HD),
        ('TOPPADDING', (0,last), (-1,last), 6),
        ('BOTTOMPADDING', (0,last), (-1,last), 6),
    ]))
    return tbl, total_valor, total_taxa

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  required=True)
    ap.add_argument("--output", required=True, help="Caminho do PDF de saída")
    ap.add_argument("--empresa", default="CARE SA")
    args = ap.parse_args()

    wb_src = openpyxl.load_workbook(args.input, data_only=True)
    cranio_tab, coluna_tab = detect_tabs(wb_src)
    print(f"Crânio  -> {cranio_tab}")
    print(f"Coluna  -> {coluna_tab}")

    cranio_all = read_sheet(wb_src[cranio_tab], args.empresa, "cranio") if cranio_tab and cranio_tab in wb_src.sheetnames else []
    coluna_all = read_sheet(wb_src[coluna_tab], args.empresa, "coluna") if coluna_tab and coluna_tab in wb_src.sheetnames else []

    by_medico = defaultdict(lambda: {"cranio": [], "coluna": []})
    for r in cranio_all:
        if r["medico"]: by_medico[normalize_key(r["medico"])]["cranio"].append(r)
    for r in coluna_all:
        if r["medico"]: by_medico[normalize_key(r["medico"])]["coluna"].append(r)

    out_dir = os.path.dirname(args.output)
    if out_dir: os.makedirs(out_dir, exist_ok=True)
    doc = SimpleDocTemplate(
        args.output, pagesize=landscape(A4),
        leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm,
        title=f"Tirinhas {args.empresa}",
    )

    print(f"\nGerando PDF com {len(by_medico)} blocos: {args.output}\n")

    def canonical_name(rows_c, rows_col):
        rows = rows_c or rows_col
        return rows[0]["medico"] if rows else "Desconhecido"

    story = []
    total_geral = 0.0; total_taxa_geral = 0.0
    for norm_name, data in sorted(by_medico.items()):
        name = canonical_name(data["cranio"], data["coluna"])
        tbl, valor, taxa = build_doctor_table(name, data["cranio"], data["coluna"], args.empresa)
        if tbl is None: continue
        total_geral += valor; total_taxa_geral += taxa
        story.append(KeepTogether(tbl))
        story.append(Spacer(1, 8))
        cr = len(data["cranio"]); co = len(data["coluna"])
        tag = (f"Cr:{cr} " if cr else "") + (f"Co:{co}" if co else "")
        flag = " *" if doctor_has_override(name) else ""
        print(f"  {normalize_name(name):<48} {tag:<10} R$ {valor:>12,.2f}  Taxa R$ {taxa:>10,.2f}{flag}")

    doc.build(story)

    print(f"\n{'─'*88}")
    print(f"  Empresa    : {args.empresa}")
    print(f"  Médicos    : {len(by_medico)}")
    print(f"  Volume     : R$ {total_geral:>12,.2f}")
    print(f"  Taxa total : R$ {total_taxa_geral:>12,.2f}")
    print(f"  PDF        : {args.output}")

if __name__ == "__main__":
    main()
